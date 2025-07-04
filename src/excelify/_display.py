from __future__ import annotations

import csv
import json
import pickle
import uuid
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Callable, Mapping, Sequence

import openpyxl
from lark import Lark

from excelify._cell_mapping import CellMapping
from excelify._col_conversion import alpha_to_int, int_to_alpha
from excelify._styler import SheetStyler
from excelify._types import Pos
from excelify.formula._parser import create_parser

if TYPE_CHECKING:
    from excelify._excelframe import ExcelFrame


DATA_FILE = ".excelify-data/data.pickle"


@dataclass
class DisplayData:
    dfs: Sequence[tuple[ExcelFrame, tuple[int, int]]]
    sheet_styler: SheetStyler


def display(
    dfs: Sequence[tuple[ExcelFrame, tuple[int, int]]], sheet_styler: SheetStyler
):
    data_path = Path(DATA_FILE)
    data_path.parent.mkdir(exist_ok=True)

    with data_path.open("wb") as f:
        pickle.dump(DisplayData(dfs=dfs, sheet_styler=sheet_styler), f)


@dataclass
class TableConfig:
    start_row: int
    start_col: int
    width: int
    height: int

    @classmethod
    def of_json(cls, d_: dict):
        return cls(
            start_row=d_["start_row"],
            start_col=d_["start_col"],
            width=d_["width"],
            height=d_["height"],
        )


def to_excel(
    dfs: Sequence[tuple[ExcelFrame, tuple[int, int]]],
    path: Path,
    *,
    index_path: Path | None = None,
) -> None:
    """Writes possibly more than one ExcelFrames to an excel file to a given `path`.
    Specify `index_path` to write down the location index of the ExcelFrames to use it
    for loading them from the .xlsx file in the future.

    Arguments:
        dfs: Sequence of ExcelFrames and its starting positions, zero-indexed.
        path: Path of an .xlsx file
        index_path: Path of an index JSON file

    """
    path = Path(path) if isinstance(path, str) else path
    mapping = CellMapping(dfs)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    assert worksheet is not None

    for df, (start_row, start_col) in dfs:
        for i, col in enumerate(df.columns):
            cells = df[col]
            worksheet.cell(row=start_row + 1, column=start_col + i + 1).value = col  # type: ignore
            start_offset = 1

            for j, cell in enumerate(cells):
                formula = cell.to_formula(mapping)
                if not cell.cell_expr.is_primitive():
                    formula = f"={formula}"
                worksheet.cell(
                    row=start_row + j + start_offset + 1, column=start_col + i + 1
                ).value = formula  # type: ignore

    workbook.save(path)

    if index_path is not None:
        index_json = [
            TableConfig(
                start_row=start_row,
                start_col=start_col,
                width=df.width,
                height=df.height,
            )
            for (df, (start_row, start_col)) in dfs
        ]
        with Path(index_path).open("w") as f:
            json.dump([asdict(config) for config in index_json], f, indent=4)


def _create_empty_dfs(
    workbook: openpyxl.Workbook, table_configs: Sequence[TableConfig]
) -> Sequence[ExcelFrame]:
    from excelify._excelframe import ExcelFrame

    worksheet = workbook.active
    assert worksheet is not None
    df_columns = [
        [
            worksheet[f"{int_to_alpha(col)}{config.start_row + 1}"].value
            for col in range(config.start_col, config.start_col + config.width)
        ]
        for config in table_configs
    ]
    return [
        ExcelFrame.empty(columns=columns, height=config.height)
        for columns, config in zip(df_columns, table_configs, strict=True)
    ]


def _get_cellpos_to_cellref_fn(
    table_configs: Sequence[TableConfig], dfs: Sequence[ExcelFrame]
) -> Callable[[str, int], tuple[ExcelFrame, int, int]]:
    def cellpos_to_cellref(column_str: str, row_idx: int):
        abs_col_idx = alpha_to_int(column_str)
        abs_row_idx = row_idx - 1

        for df, config in zip(dfs, table_configs, strict=True):
            rel_col_idx = abs_col_idx - config.start_col
            rel_row_idx = abs_row_idx - config.start_row - 1

            if (
                rel_col_idx >= 0
                and rel_col_idx < df.width
                and rel_row_idx >= 0
                and rel_row_idx < df.height
            ):
                return (df, rel_col_idx, rel_row_idx)
        raise ValueError(
            "The following cell position can't be mapped to one of the "
            f"loaded ExcelFrame's: {column_str}{row_idx}."
        )

    return cellpos_to_cellref


def _populate_df(
    workbook: openpyxl.Workbook, df: ExcelFrame, config: TableConfig, parser: Lark
) -> None:
    worksheet = workbook.active
    assert worksheet is not None
    for row_idx in range(config.height):
        for col_idx in range(config.width):
            df[df.columns[col_idx]][row_idx].cell_expr = parser.parse(
                str(
                    worksheet.cell(
                        row=config.start_row + 1 + 1 + row_idx,
                        column=config.start_col + 1 + col_idx,
                    ).value
                ).upper()
            )


def of_excel(*, path: Path | str, index_path: Path | str) -> Sequence[ExcelFrame]:
    """Loads possibly more than one ExcelFrames from an .xlsx file from a given `path`
    and `index_path`.

    Arguments:
        path: .xlsx file Path to load from
        index_path: Index JSON file path that specifies the location of the df's to load.

    Returns:
        A sequence of ExcelFrame, ordered based on the index path.
    """
    path = Path(path) if isinstance(path, str) else path
    index_path = Path(index_path) if isinstance(index_path, str) else index_path
    with index_path.open("r") as f:
        index_json = json.load(f)
    table_configs = [TableConfig.of_json(d_) for d_ in index_json]
    workbook = openpyxl.load_workbook(path, read_only=True)
    dfs = _create_empty_dfs(workbook, table_configs)

    cellpos_to_cellref = _get_cellpos_to_cellref_fn(table_configs, dfs)
    parser = create_parser(cellpos_to_cellref)

    for df, config in zip(dfs, table_configs, strict=True):
        _populate_df(workbook, df, config, parser)
    return dfs


def _try_converting_type(value: str):
    try:
        return float(value)
    except ValueError:
        return value


def of_csv(path: Path | str) -> ExcelFrame:
    """Loads an ExcelFrame from a csv file.

    Arguments:
        path: CSV file path

    Returns:
        An ExcelFrame
    """
    from excelify._excelframe import ExcelFrame

    path = Path(path) if isinstance(path, str) else path
    with path.open("r", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError(
                f"Field names could not be inferred from the csv file {path}"
            )
        columns = reader.fieldnames
        d_ = {col: [] for col in columns}
        for row in reader:
            for col in columns:
                value = _try_converting_type(row[col])
                d_[col].append(value)
        return ExcelFrame(d_)


def to_json(
    dfs: Sequence[tuple[ExcelFrame, tuple[int, int]]],
    include_header: bool = True,
    sheet_styler: SheetStyler | None = None,
):
    cell_mapping = CellMapping(dfs, header_in_table=include_header)
    tables = [
        df._to_json(start_pos=Pos.of_tuple(start_pos), cell_mapping=cell_mapping)
        for df, start_pos in dfs
    ]
    col_styles_json: dict
    if sheet_styler is not None:
        col_styles_json = sheet_styler.to_json()
    else:
        col_styles_json = {}
    return {"tables": tables, "colStyles": col_styles_json}
