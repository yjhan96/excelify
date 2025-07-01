"""ExcelFrame and functions applied to it."""

from __future__ import annotations

import copy
import uuid
from dataclasses import asdict
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence, overload
from typing_extensions import Self

import openpyxl
from tabulate import tabulate

from excelify._cell import Cell
from excelify._cell_expr import CellExpr, Constant, Empty
from excelify._cell_mapping import CellMapping, int_to_alpha
from excelify._column import Column
from excelify._element import Element
from excelify._expr import Expr
from excelify._html import NotebookFormatter
from excelify._styler import DisplayAxis, TableStyler
from excelify._types import Dimension, Pos, RawInput


def _topological_sort(cells: list[Cell]) -> list[Cell]:
    visited = set()
    result = []

    def _sort_deps(cell: Cell, result, visited):
        element = cell.element
        visited.add(element)
        for dep in cell.dependencies:
            if dep.element not in visited:
                _sort_deps(dep, result, visited)
        result.append(cell)

    for cell in cells:
        element = cell.element
        if element not in visited:
            _sort_deps(cell, result, visited)

    return result


class ExcelFrame:
    """DataFrame-like object that allows referencing other cell's values lazily
    like an Excel spreadsheet.

    Example:
        ```python
        >>> import excelify as el
        >>> df = el.ExcelFrame.empty(columns=["x", "y"], height=2)
        >>> df
        shape: (2, 2)
        +---+-------+-------+
        |   | x (A) | y (B) |
        +---+-------+-------+
        | 1 |       |       |
        | 2 |       |       |
        +---+-------+-------+
        >>> df = df.with_columns(
        ...     el.lit([0, 1]).alias("x"),
        ...     el.col("x").alias("y"),
        ... )
        >>> df
        shape: (2, 2)
        +---+-------+-------+
        |   | x (A) | y (B) |
        +---+-------+-------+
        | 1 |   0   |  A1   |
        | 2 |   1   |  A2   |
        +---+-------+-------+
        >>> df.evaluate()
        shape: (2, 2)
        +---+-------+-------+
        |   | x (A) | y (B) |
        +---+-------+-------+
        | 1 |   0   |   0   |
        | 2 |   1   |   1   |
        +---+-------+-------+

        ```
    """

    # TODO: Ideally, I'd like potentially multiple constructors implemented
    # separately, but there's no clean way to do this in Python. If I re-implement
    # the backend in C++, that should be doable.
    def __init__(
        self,
        input: Mapping[str, Iterable[RawInput | Cell | CellExpr]],
        *,
        ordered_columns: Sequence[str] | None = None,
        styler: TableStyler | None = None,
    ):
        self._id = uuid.uuid4()
        prev_cells = self._get_cell_elements(input)
        self._input: dict[str, Column] = {
            key: Column(self._id, key, values) for key, values in input.items()
        }
        prev_refs = {
            element: self._input[col_name][i]
            for (col_name, i), element in prev_cells.items()
        }
        self._update_self_refs(prev_refs)
        self._ordered_columns: list[str]
        if ordered_columns is not None:
            self._ordered_columns = copy.copy(list(ordered_columns))
        else:
            self._ordered_columns = list(self._input.keys())
        if styler:
            self._styler = styler
        else:
            self._styler = TableStyler()

    @property
    def style(self) -> TableStyler:
        return self._styler

    def _get_cell_elements(
        self,
        input: Mapping[str, Iterable[RawInput | Cell | CellExpr]],
    ) -> Mapping[tuple[str, int], Element]:
        res = {}
        for col_name, values in input.items():
            for i, value in enumerate(values):
                if isinstance(value, Cell):
                    res[(col_name, i)] = value.element
        return res

    def _update_self_refs(self, prev_refs: Mapping[Element, Cell]) -> None:
        for cells in self._input.values():
            for cell in cells:
                cell.update_cell_refs(prev_refs)

    @classmethod
    def empty(cls, *, columns: Sequence[str], height: int) -> ExcelFrame:
        """Creates an empty table with given columns and height.

        Arguments:
            columns: list of columns
            height: height of the table

        Returns:
            An empty ExcelFrame
        """
        input = {col_name: [Empty() for _ in range(height)] for col_name in columns}
        return ExcelFrame(input, ordered_columns=columns)

    @overload
    def __getitem__(self, idx_or_column: int) -> Iterable[Cell]: ...

    @overload
    def __getitem__(self, idx_or_column: str) -> Column: ...

    def __getitem__(self, idx_or_column: int | str) -> Iterable[Cell] | Column:
        if isinstance(idx_or_column, int):
            # TODO: Come back to this.
            idx = idx_or_column
            return [series[idx] for series in self._input.values()]
        else:
            column = idx_or_column
            return self._input[column]

    @property
    def height(self) -> int:
        """Height of the table."""
        if self.width == 0:
            return 0
        return len(next(iter(self._input.values())))

    @property
    def width(self) -> int:
        """Width of the table."""
        return len(self._input)

    @property
    def shape(self) -> tuple[int, int]:
        """Shape of the table."""
        return (self.height, self.width)

    @property
    def columns(self) -> list[str]:
        """List of column names in order."""
        return self._ordered_columns

    @property
    def id(self) -> uuid.UUID:
        return self._id

    def _repr_html_(self):
        return "".join(NotebookFormatter(self).render())

    def to_excel(
        self,
        path: Path | str,
        *,
        start_pos: tuple[int, int] = (0, 0),
    ) -> None:
        """Writes the ExcelFrame to `path` in an .xlsx format.

        Arguments:
            path: Path to write an .xlsx file to
            start_pos: Starting position of the table. It represents the
                table's upper left cell position.

        Returns:
            None, but the file will be written to the `path`.
        """
        path = Path(path) if isinstance(path, str) else path
        start_row, start_col = start_pos
        mapping = CellMapping([(self, (start_row, start_col))])
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        for i, col in enumerate(self._ordered_columns):
            adjusted_start_row, adjusted_start_col = start_row + 1, start_col + 1
            cells = self._input[col]
            worksheet.cell(
                row=adjusted_start_row, column=adjusted_start_col + i
            ).value = col
            start_offset = 1
            for j, cell in enumerate(cells):
                formula = cell.to_formula(mapping)
                if not cell.cell_expr.is_primitive():
                    formula = f"={formula}"
                worksheet.cell(
                    row=adjusted_start_row + j + start_offset,
                    column=adjusted_start_col + i,
                ).value = formula

        workbook.save(path)

    def _copy(self) -> ExcelFrame:
        return ExcelFrame(self._input, ordered_columns=self._ordered_columns)

    def with_columns(self, *exprs: Expr, **kwargs) -> Self:
        """Adds or modifies an expression of the column to the table and returns a new ExcelFrame.

        Example:
            ```pycon
            >>> import excelify as el
            >>> df = el.ExcelFrame({"x": [1, 2]})
            >>> df = df.with_columns(
            ...     el.map(
            ...         lambda idx: el.col("x")
            ...         if idx == 0
            ...         else el.col("cumulative_x_sum").prev(1) + el.col("x")
            ...     ).alias("cumulative_x_sum"),
            ...     x_times_two=el.col("x") * 2
            ... )
            >>> df.select(["x", "x_times_two", "cumulative_x_sum"])
            shape: (2, 3)
            +---+-------+-----------------+----------------------+
            |   | x (A) | x_times_two (B) | cumulative_x_sum (C) |
            +---+-------+-----------------+----------------------+
            | 1 |   1   |    (A1 * 2)     |          A1          |
            | 2 |   2   |    (A2 * 2)     |      (C1 + A2)       |
            +---+-------+-----------------+----------------------+

            ```

        Arguments:
            *exprs: expressions to add to the ExcelFrame
            **kwargs: Column expressions whose name will be determined by the argument name.

        Returns:
            An ExcelFrame with added/updated columns based on the passed expressions.
        """
        copy = self
        height = copy.height
        expr_list = list(exprs) + [
            expr.alias(col_name) for col_name, expr in kwargs.items()
        ]
        for expr in expr_list:
            col_name = str(expr)
            if col_name not in copy._input:
                # Create an empty column first so the column expressions can
                # refer to itself.
                copy._input[col_name] = Column(
                    copy._id, col_name, [Empty() for _ in range(height)]
                )
                copy._ordered_columns.append(col_name)
            for i, cell in enumerate(copy._input[col_name]):
                cell.set_expr(expr.get_cell_expr(copy, i))
        return copy

    def evaluate(self, inherit_style: bool = False) -> ExcelFrame:
        """Evaluate the table and return a new table with numerical values only.

        Example:
            ```pycon
            >>> import excelify as el
            >>> df = el.ExcelFrame.empty(columns=["x", "x_times_two"], height=2)
            >>> df = df.with_columns(
            ...         el.lit([1, 2]).alias("x"),
            ...         (el.col("x") * 2).alias("x_times_two"),
            ...     )
            >>> df.evaluate()
            shape: (2, 2)
            +---+-------+-----------------+
            |   | x (A) | x_times_two (B) |
            +---+-------+-----------------+
            | 1 |   1   |        2        |
            | 2 |   2   |        4        |
            +---+-------+-----------------+

            ```

        Arguments:
            inherit_style: If True, the returned table will also inherit the style
                of the original table.

        Returns:
            A new ExcelFrame where each cell represents a computed value.
        """
        cells = [cell for cells in self._input.values() for cell in cells]

        sorted_cells = _topological_sort(cells)
        for cell in sorted_cells:
            cell.compute()

        uid = uuid.uuid4()

        return ExcelFrame(
            {
                col_name: [
                    # TODO: We annoyingly create a dummy cell because
                    # we want to pass attributes. This cell gets recreated
                    # during the constructor, so it's a bit wasteful.
                    Cell(
                        Element(uid, col_name, idx),
                        Constant(value.last_value),
                        attributes=value.attributes,
                    )
                    for idx, value in enumerate(values)
                ]
                for col_name, values in self._input.items()
            },
            styler=self._styler if inherit_style else None,
        )

    def transpose(
        self,
        *,
        include_header: bool = False,
        header_name: str = "column",
        column_names: Iterable[str] | None = None,
    ) -> ExcelFrame:
        """Transpose the table.

        Example:
            ```pycon
            >>> import excelify as el
            >>> df = el.ExcelFrame({"x": [1, 2], "y": [3, 4]})
            >>> df = df.with_columns((el.col("x") + el.col("y")).alias("x_plus_y"))
            >>> df
            shape: (2, 3)
            +---+-------+-------+--------------+
            |   | x (A) | y (B) | x_plus_y (C) |
            +---+-------+-------+--------------+
            | 1 |   1   |   3   |  (A1 + B1)   |
            | 2 |   2   |   4   |  (A2 + B2)   |
            +---+-------+-------+--------------+
            >>> df.transpose(include_header=True)
            shape: (3, 3)
            +---+------------+--------------+--------------+
            |   | column (A) | column_0 (B) | column_1 (C) |
            +---+------------+--------------+--------------+
            | 1 |     x      |      1       |      2       |
            | 2 |     y      |      3       |      4       |
            | 3 |  x_plus_y  |  (B1 + B2)   |  (C1 + C2)   |
            +---+------------+--------------+--------------+

            ```

        Arguments:
            include_header: Add header as a separate column if set to True.
            header_name: Name of the header column if `include_header` is set to True.
            column_names: Name of the new ExcelFrame's columns.

        Returns:
            A transposed ExcelFrame
        """
        copy = self._copy()
        columns = []
        if include_header:
            columns.append((header_name, self.columns))

        if column_names is None:
            column_names = []

        iterator = iter(column_names)
        for i in range(self.height):
            col_name: str
            try:
                col_name = next(iterator)
            except StopIteration:
                col_name = f"column_{i}"
            columns.append((col_name, copy[i]))
        return ExcelFrame(dict(columns))

    def select(self, columns: list[str]) -> Self:
        """Select a list of columns. It can also be used to reorder the columns.

        Example:
            ```pycon
            >>> import excelify as el
            >>> df = el.ExcelFrame({"x": [1, 2], "y": [4, 5]})
            >>> df = df.with_columns((el.col("x") + el.col("y")).alias("z"))
            >>> df.select(["y", "x", "z"])
            shape: (2, 3)
            +---+-------+-------+-----------+
            |   | y (A) | x (B) |   z (C)   |
            +---+-------+-------+-----------+
            | 1 |   4   |   1   | (B1 + A1) |
            | 2 |   5   |   2   | (B2 + A2) |
            +---+-------+-------+-----------+
            >>> df.select(["x", "z"])
            shape: (2, 2)
            +---+-------+----------------+
            |   | x (A) |     z (B)      |
            +---+-------+----------------+
            | 1 |   1   | (A1 + ???:y:0) |
            | 2 |   2   | (A2 + ???:y:1) |
            +---+-------+----------------+

            ```
        """
        # TODO: Validate early on whether dropping certain column is okay or not.
        res = self
        res._ordered_columns = copy.copy(columns)
        # Update input by only taking the data specified in the column.
        res._input = {col: res._input[col] for col in res._ordered_columns}
        return res

    def _get_display_column_indices(self) -> dict[str, int]:
        column_groups = self.style.column_groups
        included_columns = [col for _, columns in column_groups for col in columns]
        remaining_columns = [col for col in self.columns if col not in included_columns]
        counter = 0
        res = {}
        for _, columns in column_groups:
            counter += 1
            for col in columns:
                res[col] = counter
                counter += 1

        return res | {c: i + counter for i, c in enumerate(remaining_columns)}

    def _get_cell_index_from_display_pos(
        self, pos: tuple[int, int], start_pos: tuple[int, int]
    ):
        pos_resolved: Pos = Pos.of_tuple(pos)
        start_pos_resolved: Pos = Pos.of_tuple(start_pos)
        col_to_idx = {c: i for i, c in enumerate(self.columns)}
        displayed_idx_to_col = {
            i: c for c, i in self._get_display_column_indices().items()
        }
        rel_pos: Pos
        match self.style.display_axis:
            case DisplayAxis.VERTICAL:
                rel_pos = Pos(
                    pos_resolved.row - start_pos_resolved.row - 1,
                    pos_resolved.col - start_pos_resolved.col,
                )
            case DisplayAxis.HORIZONTAL:
                rel_pos = Pos(
                    pos_resolved.col - start_pos_resolved.col - 1,
                    pos_resolved.row - start_pos_resolved.row,
                )
            case other:
                raise ValueError(f"Unknown axis: {other}")

        if (
            rel_pos.row >= 0
            and rel_pos.row < self.height
            and rel_pos.col in displayed_idx_to_col
        ):
            return (rel_pos.row, col_to_idx[displayed_idx_to_col[rel_pos.col]])
        else:
            return None

    def _get_display_columns(self) -> tuple[list[tuple[str, list[str]]], list[str]]:
        column_groups = self.style.column_groups
        included_columns = {col for _, columns in column_groups for col in columns}
        remaining_columns = [col for col in self.columns if col not in included_columns]
        return column_groups, remaining_columns

    def _get_col_group_column(self, group_name: str) -> list[dict]:
        return [
            {
                "formula": group_name,
                "value": group_name,
                "depIndices": [],
                "is_editable": False,
                "isPrimitive": False,
            }
        ] + [
            {"formula": "", "value": "", "depIndices": [], "is_editable": False}
            for _ in range(self.height)
        ]

    def _get_display_column(
        self, col_name: str, evaluated_df: ExcelFrame, cell_mapping: CellMapping
    ) -> list[dict]:
        display_col_name = self.style.column_renames.get(col_name, col_name)
        res = [
            {
                "formula": display_col_name,
                "value": display_col_name,
                "depIndices": [],
                "is_editable": False,
                "isPrimitive": False,
            }
        ]

        formula_column = self[col_name]
        value_column = evaluated_df[col_name]

        for formula_cell, value_cell in zip(formula_column, value_column, strict=True):
            deps = formula_cell.dependencies
            dep_indices = [cell_mapping.get_cell_index(dep.element) for dep in deps]
            formula_value = formula_cell.to_formula(cell_mapping)
            formatted_value = self.style.format_value(
                value_cell, value_cell.to_formula(cell_mapping)
            )

            res.append(
                {
                    "formula": formula_value,
                    "value": formatted_value.value,
                    "depIndices": dep_indices,
                    "is_editable": formula_cell.is_editable,
                    "color": formatted_value.color,
                }
            )
        return res

    def _to_json_vertically(self, cell_mapping: CellMapping):
        table = []
        evaluated_df = self.evaluate(inherit_style=True)
        column_groups, remaining_columns = self._get_display_columns()

        for group_name, columns in column_groups:
            col_groum_column = self._get_col_group_column(group_name)
            display_columns = [
                self._get_display_column(col_name, evaluated_df, cell_mapping)
                for col_name in columns
            ]
            table.append(col_groum_column)
            table.extend(display_columns)

        remaining_display_columns = [
            self._get_display_column(col_name, evaluated_df, cell_mapping)
            for col_name in remaining_columns
        ]
        table.extend(remaining_display_columns)
        return table

    def _get_table_json(self, cell_mapping: CellMapping) -> list[list[dict[str, Any]]]:
        match self.style.display_axis:
            case DisplayAxis.VERTICAL:
                return self._to_json_vertically(cell_mapping=cell_mapping)
            case DisplayAxis.HORIZONTAL:
                vertical_table = self._to_json_vertically(cell_mapping=cell_mapping)
                return [
                    [vertical_table[row][col] for row in range(len(vertical_table))]
                    for col in range(len(vertical_table[0]))
                ]
            case other:
                raise ValueError(f"Unknown display axis: {other}")

    def _get_display_dimension(self, table) -> Dimension:
        return Dimension(width=len(table), height=len(table[0]))

    def _to_json(self, *, start_pos: Pos, cell_mapping: CellMapping):
        table = self._get_table_json(cell_mapping)
        display_dimension = self._get_display_dimension(table)
        return {
            "table": table,
            "startPos": asdict(start_pos),
            "displayDimension": asdict(display_dimension),
        }

    def to_json(
        self, *, include_header: bool = False, start_pos: tuple[int, int] = (0, 0)
    ) -> Any:
        """Returns a JSON that represents the ExcelFrame table. excelify-app
        uses it to get the JSON-formatted state of the ExcelFrame table.

        Arguments:
            include_header: Include header in the beginning of the row if set to True
            start_pos: Starting position of the table. It represents the
                table's upper left cell position.

        Returns:
            a dict that represents JSON.
        """
        start_row, start_col = start_pos
        if include_header:
            start_row = start_row + 1
        cell_mapping = CellMapping([(self, (start_row, start_col))])
        return self._to_json(
            start_pos=Pos.of_tuple(start_pos), cell_mapping=cell_mapping
        )

    def as_str(self) -> str:
        cell_mapping = CellMapping([(self, (0, 0))], header_in_table=False)
        table = [
            [str(i + 1)]
            + [
                cell.to_formula(cell_mapping, raise_if_missing=False)
                for cell in self[i]
            ]
            for i in range(self.height)
        ]
        headers = [f"{col} ({int_to_alpha(i)})" for i, col in enumerate(self.columns)]
        return tabulate(table, headers=headers, tablefmt="pretty")

    def __repr__(self) -> str:
        shape_str = f"shape: ({self.height}, {self.width})"
        df_str = self.as_str()
        return f"""
{shape_str}
{df_str}""".strip()


def concat(dfs: Iterable[ExcelFrame]) -> ExcelFrame:
    """Concatenates ExcelFrames along the rows into one. The cell references across
    the ExcelFrames will be converted to refer to cells within the outputted
    ExcelFrame. Cell reference outside the given ExcelFrames will still be kept as is.

    Example:
        ```pycon
        >>> import excelify as el
        >>> df1 = el.ExcelFrame.empty(columns=["x", "y"], height=2)
        >>> df1 = df1.with_columns(
        ...     el.lit([1, 2]).alias("x"),
        ...     el.lit([3, 4]).alias("y"),
        ... )
        >>> df2 = el.ExcelFrame.empty(columns=["x", "y"], height=2)
        >>> df2 = df2.with_columns(
        ...     el.col("x", from_=df1).alias("x"),
        ...     el.col("y", from_=df1).alias("y"),
        ... )
        >>> el.concat([df1, df2])
        shape: (4, 2)
        +---+-------+-------+
        |   | x (A) | y (B) |
        +---+-------+-------+
        | 1 |   1   |   3   |
        | 2 |   2   |   4   |
        | 3 |  A1   |  B1   |
        | 4 |  A2   |  B2   |
        +---+-------+-------+

        ```

    Arguments:
        dfs: An iterable of ExcelFrames to concatenate

    Returns:
        A concatenated ExcelFrame

    """
    input = {}
    for i, df in enumerate(dfs):
        if i == 0:
            input = {
                col_name: [cell for cell in cells]
                for col_name, cells in df._input.items()
            }
        else:
            for col in input:
                input[col].extend(df[col])

    return ExcelFrame(input)
